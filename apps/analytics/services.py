"""
Enterprise Reporting Engine — Export Service
Handles Excel, CSV, PDF generation using openpyxl, pandas, and weasyprint.
"""
import io
import csv
import json
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any

from django.db.models import QuerySet, Sum, Count, Avg, Q
from django.http import HttpResponse
from django.utils import timezone
from django.template.loader import render_to_string


# ── Module → QuerySet registry ────────────────────────────────────────────────

def get_queryset_for_module(module: str, company_id: int, filters: dict = None) -> QuerySet:
    """Return a filtered queryset for the given module slug."""
    filters = filters or {}

    def _apply(qs, f):
        """Apply basic filter dict to a queryset."""
        for key, val in f.items():
            if val not in (None, '', [], {}):
                qs = qs.filter(**{key: val})
        return qs

    if module == 'sales_orders':
        from apps.sales.models import SalesOrder
        qs = SalesOrder.objects.filter(company_id=company_id, is_deleted=False).select_related('customer')
    elif module == 'invoices':
        from apps.sales.models import Invoice
        qs = Invoice.objects.filter(company_id=company_id, is_deleted=False).select_related('customer')
    elif module == 'customers':
        from apps.crm.models import Customer
        qs = Customer.objects.filter(company_id=company_id, is_deleted=False)
    elif module == 'leads':
        from apps.crm.models import Lead
        qs = Lead.objects.filter(company_id=company_id, is_deleted=False)
    elif module == 'purchase_orders':
        from apps.purchase.models import PurchaseOrder
        qs = PurchaseOrder.objects.filter(company_id=company_id, is_deleted=False).select_related('vendor')
    elif module == 'purchase_invoices':
        from apps.purchase.models import VendorBill
        qs = VendorBill.objects.filter(company_id=company_id, is_deleted=False).select_related('vendor')
    elif module == 'vendors':
        from apps.purchase.models import Vendor
        qs = Vendor.objects.filter(company_id=company_id, is_deleted=False)
    elif module == 'inventory':
        from apps.inventory.models import StockRecord
        qs = StockRecord.objects.filter(company_id=company_id, is_deleted=False).select_related('product', 'warehouse')
    elif module == 'products':
        from apps.inventory.models import Product
        qs = Product.objects.filter(company_id=company_id, is_deleted=False).select_related('category')
    elif module == 'employees':
        from apps.hrms.models import Employee
        qs = Employee.objects.filter(company_id=company_id, is_deleted=False).select_related('department')
    elif module == 'payroll':
        from apps.hrms.models import PayrollEntry
        qs = PayrollEntry.objects.filter(company_id=company_id, is_deleted=False).select_related('employee')
    elif module == 'attendance':
        from apps.hrms.models import Attendance
        qs = Attendance.objects.filter(company_id=company_id).select_related('employee')
    elif module == 'leave_requests':
        from apps.hrms.models import LeaveRequest
        qs = LeaveRequest.objects.filter(company_id=company_id, is_deleted=False).select_related('employee', 'leave_type')
    elif module == 'journal_entries':
        from apps.accounting.models import JournalEntry
        qs = JournalEntry.objects.filter(company_id=company_id, is_deleted=False)
    elif module == 'manufacturing_orders':
        from apps.manufacturing.models import ManufacturingOrder
        qs = ManufacturingOrder.objects.filter(company_id=company_id, is_deleted=False).select_related('bom')
    elif module == 'projects':
        from apps.projects.models import Project
        qs = Project.objects.filter(company_id=company_id, is_deleted=False).select_related('manager')
    elif module == 'tasks':
        from apps.projects.models import Task
        qs = Task.objects.filter(company_id=company_id, is_deleted=False).select_related('project', 'assigned_to')
    elif module == 'timesheets':
        from apps.projects.models import TimeLog
        qs = TimeLog.objects.filter(company_id=company_id, is_deleted=False).select_related('task', 'user')
    elif module == 'helpdesk_tickets':
        from apps.helpdesk.models import Ticket
        qs = Ticket.objects.filter(company_id=company_id, is_deleted=False).select_related('customer', 'assigned_to')
    elif module == 'assets':
        from apps.assets.models import Asset
        qs = Asset.objects.filter(company_id=company_id, is_deleted=False).select_related('category')
    else:
        from django.db.models import Model
        return None

    return _apply(qs, filters)


def serialize_row(obj, fields: List[str]) -> Dict[str, Any]:
    """Convert a model instance to a flat dict of requested fields."""
    row = {}
    for f in fields:
        val = obj
        for part in f.split('__'):
            try:
                val = getattr(val, part, None)
                if callable(val):
                    val = val()
            except Exception:
                val = None
        if isinstance(val, Decimal):
            val = float(val)
        elif hasattr(val, 'isoformat'):
            val = val.strftime('%Y-%m-%d')
        elif val is None:
            val = ''
        row[f] = str(val)
    return row


def get_data(module: str, company_id: int, columns: List[str],
             filters: dict = None, sort_by: str = '-created_at',
             limit: int = None) -> List[Dict[str, Any]]:
    """Main data fetcher — returns list of flat dicts."""
    qs = get_queryset_for_module(module, company_id, filters)
    if qs is None:
        return []
    try:
        if sort_by:
            qs = qs.order_by(sort_by)
        if limit:
            qs = qs[:limit]
        return [serialize_row(obj, columns) for obj in qs]
    except Exception as e:
        return []


# ── Export helpers ─────────────────────────────────────────────────────────────

def export_csv(data: List[Dict], filename: str = 'report.csv') -> HttpResponse:
    """Return an HttpResponse with CSV data attached."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    if not data:
        return response
    writer = csv.DictWriter(response, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    return response


def export_excel(data: List[Dict], filename: str = 'report.xlsx',
                 sheet_name: str = 'Report') -> HttpResponse:
    """Return an HttpResponse with Excel XLSX data attached using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header style
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        bottom=Side(style='thin'),
        top=Side(style='thin'),
    )

    if not data:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    headers = list(data[0].keys())

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace('__', ' > ').replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    for row_idx, row in enumerate(data, 2):
        row_fill = PatternFill(start_color='F0F4FA', end_color='F0F4FA', fill_type='solid') if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ''))
            cell.alignment = Alignment(vertical='center')
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

    # Auto-size columns
    for col_idx, header in enumerate(headers, 1):
        max_len = max(len(str(header)), max((len(str(row.get(header, ''))) for row in data), default=0))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Summary row at the bottom
    ws.append([])
    ws.cell(row=len(data) + 3, column=1, value=f'Total Records: {len(data)}').font = Font(bold=True)
    ws.cell(row=len(data) + 3, column=2, value=f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}').font = Font(italic=True, color='666666')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_pdf(data: List[Dict], report_name: str = 'Report',
               filename: str = 'report.pdf') -> HttpResponse:
    """Render a PDF using weasyprint from an HTML template."""
    try:
        from weasyprint import HTML, CSS

        html_content = render_to_string('analytics/pdf_report.html', {
            'report_name': report_name,
            'data': data,
            'headers': list(data[0].keys()) if data else [],
            'total_rows': len(data),
            'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M'),
        })

        pdf_bytes = HTML(string=html_content).write_pdf()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf_bytes)
        return response
    except Exception as e:
        # Fallback to simple PDF via reportlab
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                rightMargin=0.5*inch, leftMargin=0.5*inch,
                                topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                     fontSize=16, textColor=colors.HexColor('#1e3a5f'))
        story.append(Paragraph(report_name, title_style))
        story.append(Spacer(1, 0.2*inch))

        if data:
            headers = list(data[0].keys())
            table_data = [[h.replace('__', ' > ').replace('_', ' ').title() for h in headers]]
            for row in data[:500]:  # cap at 500 rows for PDF
                table_data.append([str(row.get(h, '')) for h in headers])

            col_width = (landscape(A4)[0] - inch) / max(len(headers), 1)
            t = Table(table_data, colWidths=[col_width] * len(headers), repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4FA')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(t)

        story.append(Spacer(1, 0.2*inch))
        story.append(Paragraph(
            f'Total: {len(data)} records | Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}',
            styles['Normal']
        ))

        doc.build(story)
        pdf = buffer.getvalue()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        return response


def get_pivot_data(data: List[Dict], row_field: str, col_field: str,
                   value_field: str) -> Dict:
    """Build a pivot table from flat data."""
    pivot = {}
    all_cols = set()

    for row in data:
        row_key = str(row.get(row_field, 'N/A'))
        col_key = str(row.get(col_field, 'N/A'))
        val = row.get(value_field, '0')
        try:
            val = float(val)
        except (ValueError, TypeError):
            val = 1.0  # count mode

        all_cols.add(col_key)
        if row_key not in pivot:
            pivot[row_key] = {}
        pivot[row_key][col_key] = pivot[row_key].get(col_key, 0) + val

    col_headers = sorted(list(all_cols))
    
    rows = []
    col_totals = [0] * len(col_headers)
    grand_total = 0

    for r_key, c_dict in pivot.items():
        row_vals = []
        row_total = 0
        for i, c_key in enumerate(col_headers):
            v = c_dict.get(c_key, 0)
            row_vals.append(v)
            row_total += v
            col_totals[i] += v
            grand_total += v
        rows.append({
            'label': r_key,
            'values': row_vals,
            'total': row_total
        })

    return {
        'row_field': row_field,
        'col_field': col_field,
        'col_headers': col_headers,
        'rows': sorted(rows, key=lambda x: str(x['label'])),
        'col_totals': col_totals,
        'grand_total': grand_total
    }
